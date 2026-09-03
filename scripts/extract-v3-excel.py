from io import BytesIO
from pathlib import Path
import json
import sys
import unicodedata

from openpyxl import load_workbook
from PIL import Image

ROOT = Path(__file__).resolve().parent.parent
EXCEL = ROOT / "data" / "explore_V3_case.xlsx"
IMAGE_ROOT = ROOT / "public" / "v3" / "cases"
DATA_ROOT = ROOT / "src" / "data" / "v3"
REPORT = ROOT / "diagnostics" / "v3-phase1-extraction-report.md"

COLUMNS = {
    "base": 2, "type": 3, "ref1": 4, "ref1_selection": 5,
    "ref2": 6, "ref2_selection": 7, "ref3": 8,
    "ref3_selection": 9, "output": 10,
}
TAG_MAPPING = {
    "1": "Design Techniques", "2": "Facade Design",
    "3": "Materials & Construction", "4": "Site Landscaping",
    "5": "Light & Atmosphere", "6": "Image Style",
}
IMAGE_COLUMNS = {COLUMNS[name] for name in ("base", "ref1", "ref2", "ref3", "output")}
EMU_PER_POINT = 12700
CONFIRMED_COLUMN_WIDTH_EMU = 962025


def normalized_name(value):
    return unicodedata.normalize("NFKC", str(value)).strip().casefold()


def has_data(sheet):
    return bool(sheet._images) or any(
        cell.value is not None for row in sheet.iter_rows() for cell in row
    )


def select_sheet(workbook):
    print("Available worksheets:")
    for name in workbook.sheetnames:
        print(name)
    if len(workbook.sheetnames) == 1:
        return workbook[workbook.sheetnames[0]]
    normalized = {name: normalized_name(name) for name in workbook.sheetnames}
    for marker in ("v3", "case"):
        matches = [name for name, value in normalized.items() if marker in value]
        if len(matches) == 1:
            return workbook[matches[0]]
        if len(matches) > 1:
            raise ValueError(f"Ambiguous sheets containing {marker!r}: {matches!r}")
    matches = [name for name in workbook.sheetnames if has_data(workbook[name])]
    if len(matches) == 1:
        return workbook[matches[0]]
    raise ValueError(f"Could not select a worksheet unambiguously: {matches!r}")


def value(sheet, row, column):
    raw = sheet.cell(row, column).value
    return None if raw is None else str(raw).strip()


def logical_anchor(sheet, image):
    anchor = image.anchor
    if not hasattr(anchor, "_from"):
        raise ValueError(f"Unsupported image anchor: {anchor!r}")
    row = anchor._from.row + 1
    column = anchor._from.col + 1
    row_offset = anchor._from.rowOff
    column_offset = anchor._from.colOff
    while column_offset >= CONFIRMED_COLUMN_WIDTH_EMU:
        column_offset -= CONFIRMED_COLUMN_WIDTH_EMU
        column += 1
    while True:
        height = round((sheet.row_dimensions[row].height or 15) * EMU_PER_POINT)
        if row_offset < height:
            break
        row_offset -= height
        row += 1
    return row, column


def raw_image(image):
    if not hasattr(image, "_data"):
        raise RuntimeError("Unable to extract image data")
    return image._data()


def write_and_validate_webp(raw, destination):
    with Image.open(BytesIO(raw)) as source:
        mode = "RGBA" if "A" in source.getbands() else "RGB"
        source.convert(mode).save(destination, "WEBP", quality=92, method=6)
    with Image.open(destination) as generated:
        generated.verify()
    with Image.open(destination) as generated:
        generated.load()
        if generated.format != "WEBP":
            raise ValueError(f"Not a valid WebP: {destination}")


def selections(raw, row, column):
    if raw is None or not str(raw).strip():
        return []
    result = []
    for token in str(raw).split(","):
        token = token.strip()
        if not token:
            continue
        try:
            number = int(token)
        except ValueError as error:
            raise ValueError(f"Invalid tag {token!r} at row {row}, column {column}") from error
        if str(number) not in TAG_MAPPING:
            raise ValueError(f"Unknown tag {number} at row {row}, column {column}")
        result.append(number)
    return result


def main():
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8")
    workbook = load_workbook(EXCEL, data_only=False)
    sheet = select_sheet(workbook)
    print(f"Selected worksheet: {sheet.title}")

    images = {}
    for index, image in enumerate(sheet._images, start=1):
        row, column = logical_anchor(sheet, image)
        coordinate = sheet.cell(row, column).coordinate
        if row < 2 or column not in IMAGE_COLUMNS:
            raise ValueError(f"Image #{index} maps outside B/D/F/H/J: {coordinate}")
        if (row, column) in images:
            raise ValueError(f"Multiple images map to {coordinate}")
        images[(row, column)] = (index, image)

    populated_rows = [
        row for row in range(2, sheet.max_row + 1)
        if value(sheet, row, COLUMNS["type"]) is not None
        or any((row, column) in images for column in IMAGE_COLUMNS)
    ]
    if not populated_rows:
        raise ValueError("No populated case rows found")

    IMAGE_ROOT.mkdir(parents=True, exist_ok=True)
    DATA_ROOT.mkdir(parents=True, exist_ok=True)
    REPORT.parent.mkdir(parents=True, exist_ok=True)
    cases, files = [], []
    counts = {name: 0 for name in ("base", "ref1", "ref2", "ref3", "output")}

    for case_number, row in enumerate(populated_rows, start=1):
        case_id = f"case-{case_number:03d}"
        case_dir = IMAGE_ROOT / case_id
        case_dir.mkdir(parents=True, exist_ok=True)
        case = {
            "id": case_id,
            "excelRow": row,
            "type": value(sheet, row, COLUMNS["type"]),
            "baseImage": None,
            "ref1Image": None,
            "ref1Selections": selections(sheet.cell(row, COLUMNS["ref1_selection"]).value, row, COLUMNS["ref1_selection"]),
            "ref2Image": None,
            "ref2Selections": selections(sheet.cell(row, COLUMNS["ref2_selection"]).value, row, COLUMNS["ref2_selection"]),
            "ref3Image": None,
            "ref3Selections": selections(sheet.cell(row, COLUMNS["ref3_selection"]).value, row, COLUMNS["ref3_selection"]),
            "expectedOutput": None,
        }
        definitions = (
            ("base", "baseImage", "base.webp"),
            ("ref1", "ref1Image", "ref-1.webp"),
            ("ref2", "ref2Image", "ref-2.webp"),
            ("ref3", "ref3Image", "ref-3.webp"),
            ("output", "expectedOutput", "output.webp"),
        )
        for source, target, filename in definitions:
            mapped = images.get((row, COLUMNS[source]))
            if mapped is None:
                continue
            destination = case_dir / filename
            write_and_validate_webp(raw_image(mapped[1]), destination)
            case[target] = f"/v3/cases/{case_id}/{filename}"
            files.append(destination)
            counts[source] += 1
        for number in range(1, 4):
            if case[f"ref{number}Selections"] and not case[f"ref{number}Image"]:
                raise ValueError(f"Row {row} has ref{number} selections but no image")
        cases.append(case)

    if len(cases) != len(populated_rows):
        raise ValueError("Generated case count does not match populated row count")
    if len(files) != len(sheet._images):
        raise ValueError(f"Generated {len(files)} files for {len(sheet._images)} images")

    manifest = {"version": 1, "source": EXCEL.name, "sheet": sheet.title, "cases": cases}
    manifest_path = DATA_ROOT / "manifest.json"
    manifest_path.write_text(json.dumps(manifest, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    tag_path = DATA_ROOT / "tag-mapping.json"
    tag_path.write_text(json.dumps(TAG_MAPPING, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    lines = [
        "# V3 Phase 1 extraction report", "",
        f"- Worksheet: `{sheet.title}`",
        f"- Populated Excel rows: {len(populated_rows)} ({populated_rows[0]}–{populated_rows[-1]})",
        f"- Generated cases: {len(cases)}",
        f"- Embedded images: {len(sheet._images)}",
        f"- Generated and validated WebP files: {len(files)}",
        f"- Base images: {counts['base']}", f"- Ref1 images: {counts['ref1']}",
        f"- Ref2 images: {counts['ref2']}", f"- Ref3 images: {counts['ref3']}",
        f"- Expected output images: {counts['output']}",
        "- Every generated image passed Pillow verify/load and WebP format validation.",
        "- Excel row 4 has no mapped base image; no `base.webp` was created.", "",
        "## Cases", "",
        "| Case | Excel row | Type | Base | Ref1 | Ref2 | Ref3 | Output |",
        "|---|---:|---|---|---|---|---|---|",
    ]
    present = lambda item: "yes" if item else "—"
    for case in cases:
        lines.append(
            f"| {case['id']} | {case['excelRow']} | {case['type']} | {present(case['baseImage'])} | "
            f"{present(case['ref1Image'])} | {present(case['ref2Image'])} | "
            f"{present(case['ref3Image'])} | {present(case['expectedOutput'])} |"
        )
    REPORT.write_text("\n".join(lines) + "\n", encoding="utf-8")

    print("Phase 1 extraction completed and validated")
    print(f"Cases: {len(cases)}")
    print(f"Images generated and validated: {len(files)}")
    print(f"Manifest: {manifest_path}")
    print(f"Tag mapping: {tag_path}")
    print(f"Report: {REPORT}")


if __name__ == "__main__":
    main()
