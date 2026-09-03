from collections import defaultdict
from io import BytesIO
from pathlib import Path
from zipfile import ZipFile
import sys
import xml.etree.ElementTree as ET

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from PIL import Image

from importlib.util import module_from_spec, spec_from_file_location


PROJECT_ROOT = Path(__file__).resolve().parent.parent
EXCEL_FILE = PROJECT_ROOT / "data" / "explore_V3_case.xlsx"
REPORT_FILE = PROJECT_ROOT / "diagnostics" / "v3-excel-image-mapping.md"

FIELDS = {
    2: "base",
    4: "ref1",
    6: "ref2",
    8: "ref3",
    10: "expected-output",
}

NS = {
    "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
    "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    "xdr": "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
    "pr": "http://schemas.openxmlformats.org/package/2006/relationships",
}
EMU_PER_POINT = 12700
BASE_COLUMN_WIDTH_EMU = 962025


def load_extractor_module():
    spec = spec_from_file_location(
        "extract_v3_excel", PROJECT_ROOT / "scripts" / "extract-v3-excel.py"
    )
    module = module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def normalized_position(ws, row, column, row_offset, column_offset):
    """Normalize offsets that land exactly beyond their anchor row/column."""

    while column_offset >= BASE_COLUMN_WIDTH_EMU:
        column_offset -= BASE_COLUMN_WIDTH_EMU
        column += 1

    while True:
        row_height = ws.row_dimensions[row].height or 15
        row_height_emu = round(row_height * EMU_PER_POINT)
        if row_offset < row_height_emu:
            break
        row_offset -= row_height_emu
        row += 1

    return row, column, row_offset, column_offset


def read_drawings(ws):
    records = []
    with ZipFile(EXCEL_FILE) as archive:
        drawing = ET.fromstring(archive.read("xl/drawings/drawing1.xml"))
        relationships = ET.fromstring(
            archive.read("xl/drawings/_rels/drawing1.xml.rels")
        )
        targets = {
            rel.attrib["Id"]: "xl/drawings/" + rel.attrib["Target"]
            for rel in relationships.findall("pr:Relationship", NS)
        }

        for index, anchor in enumerate(list(drawing), start=1):
            origin = anchor.find("xdr:from", NS)
            extent = anchor.find("xdr:ext", NS)
            blip = anchor.find(".//a:blip", NS)
            row = int(origin.findtext("xdr:row", namespaces=NS)) + 1
            column = int(origin.findtext("xdr:col", namespaces=NS)) + 1
            row_offset = int(origin.findtext("xdr:rowOff", namespaces=NS))
            column_offset = int(origin.findtext("xdr:colOff", namespaces=NS))
            normalized = normalized_position(
                ws, row, column, row_offset, column_offset
            )
            relationship_id = blip.attrib[f"{{{NS['r']}}}embed"]
            media_path = targets[relationship_id]
            raw = archive.read(media_path)
            with Image.open(BytesIO(raw)) as image:
                pixel_dimensions = image.size

            records.append(
                {
                    "index": index,
                    "raw_row": row,
                    "raw_column": column,
                    "row_offset": row_offset,
                    "column_offset": column_offset,
                    "row": normalized[0],
                    "column": normalized[1],
                    "remaining_row_offset": normalized[2],
                    "remaining_column_offset": normalized[3],
                    "display_dimensions": (
                        int(extent.attrib["cx"]),
                        int(extent.attrib["cy"]),
                    ),
                    "pixel_dimensions": pixel_dimensions,
                    "media_path": media_path,
                }
            )
    return records


def format_records(records):
    if not records:
        return "—"
    return "<br>".join(
        f"#{record['index']} {get_column_letter(record['raw_column'])}{record['raw_row']}"
        f" +({record['column_offset']},{record['row_offset']}) → "
        f"{get_column_letter(record['column'])}{record['row']}"
        for record in records
    )


def main():
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8")

    extractor = load_extractor_module()
    workbook = load_workbook(EXCEL_FILE, data_only=False)
    ws = extractor.select_v3_worksheet(workbook)
    records = read_drawings(ws)
    by_position = defaultdict(list)
    for record in records:
        by_position[(record["row"], record["column"])].append(record)

    lines = [
        "# V3 Excel image-to-data diagnostic",
        "",
        f"- Workbook: `{EXCEL_FILE.name}`",
        f"- Selected worksheet: `{ws.title}`",
        f"- Worksheet extent: {ws.max_row} rows × {ws.max_column} columns",
        f"- Data rows: 2–18 (17 rows)",
        f"- Embedded drawing objects: {len(records)}",
        "- No images were extracted and no manifest was generated.",
        "",
        "Offsets are OOXML EMUs. `→` shows the effective cell after consuming a full "
        "row/column offset. Row heights 2–12 are 742950 EMU (58.5 pt); the two "
        "A-column anomalies use a 962025 EMU horizontal offset, exactly matching "
        "the B-column drawing origin used elsewhere.",
        "",
        "## Row mapping",
        "",
        "| Row | Base/type value | Base B: actual anchor(s) | Ref1 D: actual anchor(s) | Ref2 F: actual anchor(s) | Ref3 H: actual anchor(s) | Output J: actual anchor(s) | Image indices | Dimensions and OOXML media filename |",
        "|---:|---|---|---|---|---|---|---|---|",
    ]

    for row in range(2, 19):
        row_records = [record for record in records if record["row"] == row]
        field_records = {
            column: by_position.get((row, column), []) for column in FIELDS
        }
        inventory = "<br>".join(
            f"#{record['index']}: {record['pixel_dimensions'][0]}×{record['pixel_dimensions'][1]} px; "
            f"display {record['display_dimensions'][0]}×{record['display_dimensions'][1]} EMU; "
            f"`{record['media_path']}`; temp: not extracted"
            for record in row_records
        ) or "—"
        indices = ", ".join(f"#{record['index']}" for record in row_records) or "—"
        type_value = ws.cell(row=row, column=3).value
        lines.append(
            f"| {row} | {type_value or '—'} | {format_records(field_records[2])} | "
            f"{format_records(field_records[4])} | {format_records(field_records[6])} | "
            f"{format_records(field_records[8])} | {format_records(field_records[10])} | "
            f"{indices} | {inventory} |"
        )

    lines.extend(
        [
            "",
            "## Rows 2–7 reasoning",
            "",
            "- **Row 2:** #16 and #17 are directly positioned at B2 and D2. #18 is directly at J2. The apparent #4/#6 duplicates have a 742950 EMU vertical offset—one complete row 2 height—so their visible top edges begin in row 3, not row 2. E2 contains ref1 selections while G2/I2 are empty, which agrees with B/D/J only.",
            "- **Row 3:** #19 is directly at B3 and #2 directly at H3. #4, #5 and #6 normalize from row-2 anchors to D3, F3 and J3 because their vertical offsets equal the full row-2 height. E3/G3/I3 all contain selections, agreeing with D/F/H reference images.",
            "- **Row 4:** #8, #3 and #20 are directly at F4, H4 and J4. #7 normalizes from D3 to D4 because its 742950 EMU offset equals the full row-3 height. E4/G4/I4 all contain selections, agreeing with D/F/H. No drawing object occupies B4 after geometric normalization, so the base image for row 4 remains unresolved.",
            "- **Row 5:** #10 is stored as A5 plus a 962025 EMU horizontal offset. Its effective x-position exactly matches the direct B-column origin of other base images, so geometry—not proximity—places it at B5. #21, #22 and #23 are directly at D5, F5 and J5. E5/G5 contain selections and I5 is empty, agreeing with D/F and no H image.",
            "- **Row 6:** #24, #25, #26 and #27 are directly positioned at B6, D6, F6 and J6. E6/G6 contain selections and I6 is empty, agreeing with the two references.",
            "- **Row 7:** #9 uses A7 plus the same exact 962025 EMU horizontal offset, placing its effective origin at B7. #28–#31 are directly positioned at D7, F7, H7 and J7. E7/G7/I7 all contain selections, agreeing with three references.",
            "",
            "## Unresolved issue",
            "",
            "After offset normalization, all duplicate/unexpected anchors are resolved geometrically. However, row 4 has no embedded image at effective B4. The workbook does not establish whether row 4 intentionally reuses another row's base image or is missing a base image. That mapping must not be inferred silently.",
        ]
    )

    REPORT_FILE.parent.mkdir(parents=True, exist_ok=True)
    REPORT_FILE.write_text("\n".join(lines) + "\n", encoding="utf-8")
    print(f"Selected worksheet: {ws.title}")
    print(f"Diagnostic report: {REPORT_FILE}")
    print(f"Drawing objects inspected: {len(records)}")
    print("Manifest generated: no")
    print("Temporary images extracted: no")
    print("Ambiguity: row 4 has no effective B4/base drawing")


if __name__ == "__main__":
    main()
