from csv import DictWriter
from io import BytesIO
from pathlib import Path
from tempfile import mkdtemp
from zipfile import ZipFile
import json
import sys
import xml.etree.ElementTree as ET

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from PIL import Image, ImageDraw, ImageFont, ImageOps


PROJECT_ROOT = Path(__file__).resolve().parent.parent
WORKBOOK_PATH = PROJECT_ROOT / "data" / "explore_V3_case.xlsx"

NS = {
    "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
    "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    "xdr": "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
    "pr": "http://schemas.openxmlformats.org/package/2006/relationships",
}


def select_worksheet(workbook):
    print("Available worksheets:")
    for name in workbook.sheetnames:
        print(name)

    if len(workbook.sheetnames) != 1:
        raise ValueError(
            "Diagnostic extraction requires an unambiguous single-sheet workbook."
        )
    return workbook[workbook.sheetnames[0]]


def drawing_records(archive):
    drawing = ET.fromstring(archive.read("xl/drawings/drawing1.xml"))
    relationships = ET.fromstring(
        archive.read("xl/drawings/_rels/drawing1.xml.rels")
    )
    targets = {
        relationship.attrib["Id"]: "xl/drawings/" + relationship.attrib["Target"]
        for relationship in relationships.findall("pr:Relationship", NS)
    }

    for index, anchor in enumerate(list(drawing), start=1):
        origin = anchor.find("xdr:from", NS)
        blip = anchor.find(".//a:blip", NS)
        row = int(origin.findtext("xdr:row", namespaces=NS)) + 1
        column = int(origin.findtext("xdr:col", namespaces=NS)) + 1
        relationship_id = blip.attrib[f"{{{NS['r']}}}embed"]
        yield {
            "image_index": index,
            "anchor_cell": f"{get_column_letter(column)}{row}",
            "row": row,
            "column": column,
            "column_letter": get_column_letter(column),
            "column_offset_emu": int(
                origin.findtext("xdr:colOff", namespaces=NS)
            ),
            "row_offset_emu": int(origin.findtext("xdr:rowOff", namespaces=NS)),
            "media_path": targets[relationship_id],
        }


def create_contact_sheet(records, image_directory, output_path):
    columns = 5
    tile_width = 260
    tile_height = 210
    padding = 12
    label_height = 34
    rows = (len(records) + columns - 1) // columns
    sheet = Image.new(
        "RGB",
        (columns * tile_width, rows * tile_height),
        (28, 28, 30),
    )
    draw = ImageDraw.Draw(sheet)
    font = ImageFont.load_default(size=16)

    for position, record in enumerate(records):
        tile_x = (position % columns) * tile_width
        tile_y = (position // columns) * tile_height
        image_path = image_directory / record["extracted_filename"]
        with Image.open(image_path) as source:
            thumbnail = ImageOps.contain(
                source.convert("RGB"),
                (tile_width - padding * 2, tile_height - label_height - padding * 2),
            )
        image_x = tile_x + (tile_width - thumbnail.width) // 2
        image_y = tile_y + padding
        sheet.paste(thumbnail, (image_x, image_y))
        label = f"#{record['image_index']}  {record['anchor_cell']}"
        draw.text(
            (tile_x + padding, tile_y + tile_height - label_height),
            label,
            fill=(245, 245, 245),
            font=font,
        )

    sheet.save(output_path, "PNG", optimize=True)


def main():
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8")

    workbook = load_workbook(WORKBOOK_PATH, data_only=False)
    worksheet = select_worksheet(workbook)
    output_directory = Path(mkdtemp(prefix="airi-v3-excel-diagnostic-"))
    image_directory = output_directory / "images"
    image_directory.mkdir()
    records = []

    with ZipFile(WORKBOOK_PATH) as archive:
        for record in drawing_records(archive):
            raw = archive.read(record["media_path"])
            with Image.open(BytesIO(raw)) as image:
                image_format = image.format or "UNKNOWN"
                width, height = image.size
                extension = Path(record["media_path"]).suffix.lower()

            filename = (
                f"image-{record['image_index']:03d}-{record['anchor_cell']}"
                f"{extension}"
            )
            (image_directory / filename).write_bytes(raw)
            records.append(
                {
                    **record,
                    "width": width,
                    "height": height,
                    "image_format": image_format,
                    "extracted_filename": filename,
                }
            )

    csv_path = output_directory / "image-mapping.csv"
    with csv_path.open("w", newline="", encoding="utf-8-sig") as csv_file:
        writer = DictWriter(csv_file, fieldnames=list(records[0]))
        writer.writeheader()
        writer.writerows(records)

    json_path = output_directory / "image-mapping.json"
    json_path.write_text(
        json.dumps(
            {
                "workbook": str(WORKBOOK_PATH),
                "worksheet": worksheet.title,
                "embedded_image_count": len(records),
                "images": records,
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )

    contact_sheet_path = output_directory / "contact-sheet.png"
    create_contact_sheet(records, image_directory, contact_sheet_path)

    print(f"Selected worksheet: {worksheet.title}")
    print(f"Images extracted: {len(records)}")
    print(f"Temporary directory: {output_directory}")
    print(f"Contact sheet: {contact_sheet_path}")
    print(f"CSV mapping: {csv_path}")
    print(f"JSON mapping: {json_path}")
    print("Production manifest generated: no")


if __name__ == "__main__":
    main()
